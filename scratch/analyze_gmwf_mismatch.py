import json
import openpyxl
from collections import defaultdict

def clean_receipt(receipt):
    if receipt is None:
        return ""
    rcpt_str = str(receipt).strip()
    if rcpt_str.endswith('.0'):
        rcpt_str = rcpt_str[:-2]
    return rcpt_str

# Load JSON records
with open('e:/GMWF/gmwf/import_donations-1.json', 'r', encoding='utf-8') as f:
    json_records = json.load(f)

json_by_exact_receipt = {r['receiptNo']: r for r in json_records}

wb = openpyxl.load_workbook('e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx', read_only=True)

# Scan for duplicate receipts across all sheets
receipt_counts = defaultdict(int)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    receipt_idx = 1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            continue
        other_fields = [row[i] for i in range(len(row)) if i != receipt_idx and row[i] is not None]
        if not other_fields:
            continue
        receipt = row[receipt_idx]
        clean_rc = clean_receipt(receipt)
        if clean_rc:
            receipt_counts[clean_rc] += 1
            
duplicate_receipts = {k for k, v in receipt_counts.items() if v > 1}
receipt_occurrences = defaultdict(int)

reasons = defaultdict(float)
reason_counts = defaultdict(int)

for sheet_name in ['GMWF', 'GMWF-25']:
    ws = wb[sheet_name]
    headers = None
    first_row_idx = -1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            first_row_idx = r_idx
            break
            
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= first_row_idx:
            continue
        if not any(x is not None for x in row):
            continue
            
        clean_rc = clean_receipt(row[1])
        if not clean_rc:
            continue
            
        receipt_no = clean_rc
        if receipt_no in duplicate_receipts:
            receipt_occurrences[receipt_no] += 1
            occ_num = receipt_occurrences[receipt_no]
            suffix = "-a" if occ_num == 1 else "-b"
            receipt_no = f"{receipt_no}{suffix}"
            
        amount_col2 = row[2]
        val_float = 0.0
        if amount_col2 is not None:
            try:
                val_float = float(amount_col2)
            except ValueError:
                pass
                
        if val_float > 0:
            if receipt_no in json_by_exact_receipt:
                json_rec = json_by_exact_receipt[receipt_no]
                json_amt = json_rec['amount']
                diff = val_float - json_amt
                if diff != 0:
                    entry_type = json_rec.get('entryType')
                    notes = json_rec.get('notes', '')
                    
                    if entry_type == 'goods':
                        cat = 'Goods (In-kind)'
                    elif 'CANCELLED' in notes:
                        cat = 'Cancelled / Zeroed'
                    else:
                        cat = 'Other Mismatch'
                        
                    reasons[cat] += diff
                    reason_counts[cat] += 1
                    
                    if cat == 'Other Mismatch' or (cat == 'Cancelled / Zeroed' and 'cancel' not in str(row[3]).lower() and 'cancel' not in str(row[2]).lower()):
                         print(f"[{sheet_name}] Row {r_idx+1}: Receipt {receipt_no} | Excel={val_float}, JSON={json_amt} | Cat={cat} | Notes: {notes}")
                         print(f"  Row details: {row[:10]}")
            else:
                reasons['Missing in JSON'] += val_float
                reason_counts['Missing in JSON'] += 1
                print(f"[{sheet_name}] Row {r_idx+1}: Receipt {receipt_no} | Excel={val_float} | Missing in JSON")
                print(f"  Row details: {row[:10]}")

print("\n=== SUMMARY OF DIFFERENCES ===")
for cat, diff in reasons.items():
    print(f"{cat}: count={reason_counts[cat]}, diff_sum={diff:,.2f}")

wb.close()
