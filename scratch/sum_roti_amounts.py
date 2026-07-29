import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    goods_keywords = ['pcs', 'kg', 'roti', 'quran']
    
    total_roti_amount = 0.0
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            continue
        donor_name = str(row[3]).strip() if (sheet_name == 'GMWF-25' and len(row) > 3 and row[3] is not None) else ""
        raw_amount = row[2] if len(row) > 2 else None
        
        name_str = str(donor_name).lower()
        amt_str = str(raw_amount or "").lower()
        
        name_is_goods = any(term in name_str for term in goods_keywords)
        amt_is_goods = any(term in amt_str for term in goods_keywords)
        
        if name_is_goods or amt_is_goods:
            if raw_amount is not None:
                try:
                    val_clean = "".join(c for c in str(raw_amount) if c.isdigit() or c=='.')
                    val = float(val_clean)
                    total_roti_amount += val
                except:
                    pass
    if total_roti_amount > 0:
        print(f"Sheet {sheet_name}: Sum of goods/roti amount column = {total_roti_amount:,}")
