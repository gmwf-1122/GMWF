import openpyxl
import datetime
import re

def extract_amount_from_row(row, main_amount, receipt_no):
    if main_amount is not None:
        try:
            val = float(main_amount)
            if val > 0:
                return val
        except ValueError:
            pass
            
    # Check specific columns that can contain amount: Name (3), Deposit Date (6 - sometimes has amount string), Deposit Receipt (7), Deposited Amount (8)
    for idx in [3, 6, 7, 8]:
        if idx >= len(row):
            continue
        cell = row[idx]
        if cell is None:
            continue
        if isinstance(cell, (datetime.datetime, datetime.date)):
            continue
            
        cell_str = str(cell).strip().upper()
        
        # Skip if it is a date-like string
        if re.match(r'^\d{4}-\d{2}-\d{2}', cell_str):
            continue
            
        # Skip names of collectors/received by and simple ONLINE/CANCELLED strings
        if cell_str in ['ZAHEER', 'MUSTANSAR', 'ALI RAZA', 'SAMEER', 'ANS', 'MUZAMIL', 'ONLINE', 'CANCELLED', 'ONLINE TRANSFER', 'ONLINE DEPOSIT']:
            continue
            
        # 1. Match '80K', '5K', '2K' etc.
        k_match = re.search(r'\b(\d+(?:\.\d+)?)\s*K\b', cell_str)
        if k_match:
            return float(k_match.group(1)) * 1000.0
            
        # 2. Match numeric values
        nums = re.findall(r'\b\d+\b', cell_str)
        for num in nums:
            val = float(num)
            # Skip if it matches the receipt number
            if val == receipt_no:
                continue
            # Avoid cell numbers or bank receipt numbers (typically 9+ digits)
            if len(num) >= 9:
                continue
            # Skip year values unless they are accompanied by other number info
            if val in [2024, 2025, 2026]:
                continue
            if val > 0:
                return val
                
    return 0.0

file_path = "e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx"
wb = openpyxl.load_workbook(file_path, read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    headers = None
    first_non_empty_row_idx = -1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            first_non_empty_row_idx = r_idx
            break
            
    receipt_idx = 1
    amount_idx = 2
    
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= first_non_empty_row_idx:
            continue
        if not any(x is not None for x in row):
            continue
            
        receipt = row[receipt_idx]
        if receipt is None:
            continue
            
        try:
            rcpt_num = float(str(receipt).strip())
        except ValueError:
            rcpt_num = 0
            
        amount = row[amount_idx]
        
        is_anomalous = False
        if amount is None:
            is_anomalous = True
        else:
            try:
                val = float(amount)
                if val == 0:
                    is_anomalous = True
            except ValueError:
                is_anomalous = True
                
        if is_anomalous:
            other_fields = [row[i] for i in range(len(row)) if i != receipt_idx and row[i] is not None]
            if not other_fields:
                continue
                
            extracted = extract_amount_from_row(row, amount, rcpt_num)
            row_str = " | ".join(str(x) for x in row if x is not None)
            is_goods = 'pcs' in row_str.lower() or 'kg' in row_str.lower() or 'roti' in row_str.lower()
            
            print(f"Row {r_idx+1}: Receipt={receipt}, OrigAmount={amount}, Extracted={extracted}, Goods={is_goods}")
            print(f"  Data: {row[:10]}")

wb.close()
