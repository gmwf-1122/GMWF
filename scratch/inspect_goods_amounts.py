import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    
    # Let's find rows matching goods_keywords
    goods_keywords = ['pcs', 'kg', 'roti', 'quran']
    
    headers = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            break
            
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            continue
        # check columns 1, 2, 3, etc
        donor_name = str(row[3]).strip() if (sheet_name == 'GMWF-25' and len(row) > 3 and row[3] is not None) else ""
        if sheet_name != 'GMWF-25':
            # col index for name is not present, we use amount col to search or notes col
            donor_name = ""
            
        raw_amount = row[2] if len(row) > 2 else None
        
        name_str = str(donor_name).lower()
        amt_str = str(raw_amount or "").lower()
        
        name_is_goods = any(term in name_str for term in goods_keywords)
        amt_is_goods = any(term in amt_str for term in goods_keywords)
        
        if name_is_goods or amt_is_goods:
            print(f"Row {r_idx+1}: donor_name='{donor_name}', raw_amount='{raw_amount}'")
