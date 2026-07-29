import openpyxl
import re

wb = openpyxl.load_workbook('e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx', read_only=True)

for sheet_name in ['MASJID', 'MASJID-25']:
    ws = wb[sheet_name]
    print(f"\n================ SHEET: {sheet_name} ================")
    
    headers = None
    first_row_idx = -1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            first_row_idx = r_idx
            break
            
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= first_row_idx:
            continue
        if not any(x is not None for x in row):
            continue
            
        receipt = row[1] if len(row) > 1 else None
        raw_amount = row[2] if len(row) > 2 else None
        
        # Mimic parse_excel_to_json.py logic
        donor_name = ""
        name_idx = -1
        received_by_idx = 3
        
        # donor_name is not populated in Masjid sheets because name_idx is -1, but let's check
        is_cancelled = False
        
        # Check if cancelled
        if 'cancel' in str(raw_amount).lower():
            is_cancelled = True
            
        # Try primary amount
        if raw_amount is not None:
            try:
                # parse_amount_str handles:
                # numeric values, rejects 2000-2100 as years, etc.
                s = str(raw_amount).strip().upper()
                if s not in ('ONLINE', 'ONLINE TRANSFER', 'ONLINE DEPOSIT'):
                    m2 = re.match(r'^([\d\.]+)', s)
                    if m2:
                        v = float(m2.group(1))
                        if not (2000 <= v <= 2100):
                            amount_val = v
                        else:
                            amount_val = None
                    else:
                        amount_val = None
                else:
                    amount_val = None
            except:
                amount_val = None
        else:
            amount_val = None
            
        # In parse_excel_to_json.py:
        # if is_cancelled: amount_val = 0.0
        # If amount_val is 0.0 or None, it is treated as cancelled
        
        is_actually_cancelled_by_script = is_cancelled or amount_val is None or amount_val == 0.0
        
        # Let's print rows where the raw amount is a positive number, but the script treated it as cancelled/zero
        if raw_amount is not None:
            try:
                raw_float = float(str(raw_amount).strip())
                if raw_float > 0 and is_actually_cancelled_by_script:
                    print(f"Row {r_idx+1}: Receipt={receipt}, RawAmount={raw_amount} (float={raw_float}) -> Treated as CANCELLED/ZERO by script!")
                    print(f"  Row details: {row[:8]}")
            except ValueError:
                pass

wb.close()
