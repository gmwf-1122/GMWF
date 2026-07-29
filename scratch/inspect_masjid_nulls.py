import openpyxl

wb = openpyxl.load_workbook('e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx', read_only=True)
ws = wb['MASJID-25']

headers = None
for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if any(x is not None for x in row):
        headers = row
        break

print("MASJID-25 headers:", headers)

count = 0
for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if r_idx == 0:
        continue
    # If any cell is populated
    if not any(x is not None for x in row):
        continue
        
    amount = row[2]
    receipt = row[1]
    
    if amount is None or str(amount).strip() == "" or amount == 0:
        count += 1
        print(f"Row {r_idx+1}: {[str(x)[:20] if x is not None else None for x in row[:10]]}")

print(f"Total null/zero amount rows in MASJID-25: {count}")
wb.close()
