import openpyxl
import datetime

wb = openpyxl.load_workbook('e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx', read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n================ SHEET: {sheet_name} ================")
    
    # Find headers
    headers = None
    first_row_idx = -1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            first_row_idx = r_idx
            break
            
    print(f"Header row index: {first_row_idx}")
    print(f"Headers: {headers}")
    
    # Calculate sum of column index 2 (Amount)
    col2_sum = 0
    non_numeric_rows = []
    zero_rows_with_data = []
    total_rows = 0
    
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= first_row_idx:
            continue
        # Check if completely empty
        if not any(x is not None for x in row):
            continue
            
        total_rows += 1
        val = row[2] if len(row) > 2 else None
        
        # Check if it has a receipt number
        receipt = row[1] if len(row) > 1 else None
        
        if val is not None:
            try:
                col2_sum += float(val)
            except (ValueError, TypeError):
                non_numeric_rows.append((r_idx + 1, receipt, val, row))
        else:
            zero_rows_with_data.append((r_idx + 1, receipt, val, row))
            
    print(f"Total non-empty data rows counted: {total_rows}")
    print(f"Sum of Column 2 (Amount): {col2_sum:,.2f}")
    print(f"Non-numeric Column 2 rows (count: {len(non_numeric_rows)}):")
    for r_num, rcpt, val, r in non_numeric_rows[:15]:
        print(f"  Row {r_num}: Receipt={rcpt}, Col2={val} | Row data: {r[:8]}")
    print(f"Null Column 2 rows (count: {len(zero_rows_with_data)}):")
    for r_num, rcpt, val, r in zero_rows_with_data[:15]:
        print(f"  Row {r_num}: Receipt={rcpt}, Col2={val} | Row data: {r[:8]}")

wb.close()
