import openpyxl
import json

def clean_receipt(receipt):
    if receipt is None:
        return ""
    rcpt_str = str(receipt).strip()
    if rcpt_str.endswith('.0'):
        rcpt_str = rcpt_str[:-2]
    return rcpt_str

# Load JSON records
with open('e:/GMWF/gmwf/import_donations-1.json', 'r', encoding='utf-8') as f:
    json_records = json.load(f)

# Keep only jamia (Masjid) records
json_jamia = [r for r in json_records if r['categoryId'] == 'jamia']
json_by_receipt = {}
for r in json_jamia:
    rc = r['receiptNo'].split('-')[0] # strip suffix like -a if any
    json_by_receipt[rc] = r

print(f"Total Jamia records in JSON: {len(json_jamia)}")

wb = openpyxl.load_workbook('e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx', read_only=True)

for sheet_name in ['MASJID', 'MASJID-25']:
    ws = wb[sheet_name]
    print(f"\nComparing sheet: {sheet_name}")
    
    headers = None
    first_row_idx = -1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            first_row_idx = r_idx
            break
            
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= first_row_idx:
            continue
        # Check if completely empty
        if not any(x is not None for x in row):
            continue
            
        receipt = clean_receipt(row[1])
        if not receipt:
            continue
            
        amount_col2 = row[2]
        # Deposited amount column is col 5 for MASJID, col 6 for MASJID-25
        dep_amount_col = 5 if sheet_name == 'MASJID' else 6
        amount_col_dep = row[dep_amount_col] if len(row) > dep_amount_col else None
        
        # Determine if this row is considered "blank" by parse_excel_to_json.py
        max_col_to_check = 4
        other_fields_populated = False
        for col_i in range(min(len(row), max_col_to_check)):
            if col_i != 1 and row[col_i] is not None:
                if str(row[col_i]).strip():
                    other_fields_populated = True
                    break
        
        if not other_fields_populated:
            # skipped as blank, so expected in JSON is nothing (should not be in JSON)
            if receipt in json_by_receipt:
                print(f"SHOULD BE SKIPPED BUT IN JSON: Receipt {receipt} (Row {r_idx+1})")
            continue
            
        # Check if this receipt is in JSON
        if receipt not in json_by_receipt:
            print(f"MISSING Receipt {receipt} (Row {r_idx+1}): Col2={amount_col2}, Col{dep_amount_col}={amount_col_dep}")
            print(f"  Row details: {row[:8]}")
        else:
            json_rec = json_by_receipt[receipt]
            json_amt = json_rec['amount']
            notes = json_rec.get('notes', '')
            
            # If cancelled in JSON, check if it was cancelled in Excel
            is_cancelled_in_json = 'CANCELLED' in notes
            
            # Expected amount
            expected = 0.0
            if not is_cancelled_in_json:
                if amount_col2 is not None:
                    try:
                        expected = float(amount_col2)
                    except ValueError:
                        pass
                if expected == 0.0 and amount_col_dep is not None:
                    try:
                        expected = float(amount_col_dep)
                    except ValueError:
                        pass
                    
            if json_amt != expected:
                print(f"AMOUNT MISMATCH Receipt {receipt} (Row {r_idx+1}): Excel Expected={expected} (Col2={amount_col2}, Col{dep_amount_col}={amount_col_dep}), JSON={json_amt} (Notes: {notes})")
                print(f"  Row details: {row[:8]}")

wb.close()
