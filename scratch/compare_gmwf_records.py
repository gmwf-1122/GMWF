import openpyxl
import json
from collections import defaultdict

def clean_receipt(receipt):
    if receipt is None:
        return ""
    rcpt_str = str(receipt).strip()
    if rcpt_str.endswith('.0'):
        rcpt_str = rcpt_str[:-2]
    return rcpt_str

# Load JSON records
with open('e:/GMWF/gmwf/import_donations-1.json', 'r', encoding='utf-8') as f:
    json_records = json.load(f)

# Map by exact receipt number
json_by_exact_receipt = {r['receiptNo']: r for r in json_records}

wb = openpyxl.load_workbook('e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx', read_only=True)

# Step 1: Scan for duplicate receipts across all sheets to identify duplicate_receipts
receipt_counts = defaultdict(int)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            break
    receipt_idx = 1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            continue
        other_fields = [row[i] for i in range(len(row)) if i != receipt_idx and row[i] is not None]
        if not other_fields:
            continue
        receipt = row[receipt_idx]
        clean_rc = clean_receipt(receipt)
        if clean_rc:
            receipt_counts[clean_rc] += 1
            
duplicate_receipts = {k for k, v in receipt_counts.items() if v > 1}
print(f"Total duplicate receipt numbers identified: {len(duplicate_receipts)}")

# Step 2: Compare each sheet row-by-row
receipt_occurrences = defaultdict(int)

for sheet_name in ['GMWF', 'GMWF-25']:
    ws = wb[sheet_name]
    print(f"\nComparing sheet: {sheet_name}")
    
    headers = None
    first_row_idx = -1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            first_row_idx = r_idx
            break
            
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= first_row_idx:
            continue
            
        # Determine if row is blank (using parser logic)
        max_col_to_check = 6 if sheet_name == 'GMWF-25' else 4
        other_fields_populated = False
        for col_i in range(min(len(row), max_col_to_check)):
            if col_i != 1 and row[col_i] is not None:
                if str(row[col_i]).strip():
                    other_fields_populated = True
                    break
        
        # Also check bank/online columns for GMWF-25
        if not other_fields_populated and sheet_name == 'GMWF-25':
            for col_i in [6, 7]:
                if col_i < len(row) and row[col_i] is not None:
                    if str(row[col_i]).strip():
                        other_fields_populated = True
                        break
                        
        if not other_fields_populated:
            continue
            
        clean_rc = clean_receipt(row[1])
        if not clean_rc:
            continue
            
        # Get final receipt number (with suffix)
        receipt_no = clean_rc
        if receipt_no in duplicate_receipts:
            receipt_occurrences[receipt_no] += 1
            occ_num = receipt_occurrences[receipt_no]
            suffix = "-a" if occ_num == 1 else "-b"
            receipt_no = f"{receipt_no}{suffix}"
        else:
            # Still track it in case it's in another category
            pass
            
        amount_col2 = row[2]
        
        # Check if this receipt is in JSON
        if receipt_no not in json_by_exact_receipt:
            # Try to see if it is in another category (e.g. jamia)
            # Find in all json records
            matching_rec = [r for r in json_records if r['receiptNo'] == receipt_no]
            if matching_rec:
                print(f"Receipt {receipt_no} in different category: {matching_rec[0]['categoryId']}")
            else:
                print(f"MISSING Receipt {receipt_no} (Row {r_idx+1}): Col2={amount_col2}")
                print(f"  Row details: {row[:10]}")
        else:
            json_rec = json_by_exact_receipt[receipt_no]
            json_amt = json_rec['amount']
            notes = json_rec.get('notes', '')
            is_cancelled_in_json = 'CANCELLED' in notes
            is_goods = json_rec.get('entryType') == 'goods'
            
            # Expected amount
            expected = 0.0
            if not is_cancelled_in_json and not is_goods:
                if amount_col2 is not None:
                    try:
                        # Mimic parse_amount_str
                        s = str(amount_col2).strip().upper()
                        # If the value is a number and not a year-like (2000-2100)
                        try:
                            v = float(amount_col2)
                            if 2000 <= v <= 2100:
                                expected = 0.0 # because parser currently skips it
                            else:
                                expected = v
                        except ValueError:
                            # Not direct float
                            expected = 0.0
                    except:
                        expected = 0.0
                        
            if json_amt != expected:
                print(f"AMOUNT MISMATCH Receipt {receipt_no} (Row {r_idx+1}): Excel Expected={expected} (Col2={amount_col2}), JSON={json_amt} (Notes: {notes})")
                print(f"  Row details: {row[:10]}")

wb.close()
