import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx", data_only=True)
sheet = wb['GMWF']

print("Row index, Donor Name, Amount Col, Other Columns...")
# Let's inspect rows to find where 19,180 comes from or what amounts are cancelled
total_amount_col = 0
cancelled_amount = 0
for row_idx in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row_idx, column=1).value
    amount = sheet.cell(row=row_idx, column=2).value
    col3 = sheet.cell(row=row_idx, column=3).value
    col4 = sheet.cell(row=row_idx, column=4).value
    
    if name is None and amount is None and col3 is None and col4 is None:
        continue
        
    # Check if this row is considered cancelled or goods by the parser
    name_str = str(name or "").lower()
    amt_str = str(amount or "").lower()
    
    is_cancelled = 'cancel' in name_str or 'cancel' in amt_str
    
    goods_keywords = ['pcs', 'kg', 'roti', 'quran']
    is_goods = any(k in name_str or k in amt_str for k in goods_keywords)
    
    # Try to parse amount
    val = 0.0
    if amount is not None:
        try:
            # strip non-numeric
            val_clean = "".join(c for c in str(amount) if c.isdigit() or c=='.')
            val = float(val_clean)
        except:
            pass
            
    if is_cancelled:
        print(f"Row {row_idx}: CANCELLED - Name: '{name}', Amount: '{amount}', val={val}")
        cancelled_amount += val
    elif is_goods:
        print(f"Row {row_idx}: GOODS - Name: '{name}', Amount: '{amount}', val={val}")
    elif val > 0:
        total_amount_col += val

print(f"Total parsed active cash/online from col 2: {total_amount_col:,}")
print(f"Total cancelled from col 2: {cancelled_amount:,}")
