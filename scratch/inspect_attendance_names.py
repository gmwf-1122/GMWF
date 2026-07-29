import openpyxl

print("=== Checking Monthly Salary File ===")
wb_sal = openpyxl.load_workbook('GMWF  Monthly Salary.xlsx', read_only=True)
sal_names = set()
for sheet in wb_sal.sheetnames:
    ws = wb_sal[sheet]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if len(row) > 2 and row[2]:
            name = str(row[2]).strip()
            sal_names.add(name)
            if 'rehman' in name.lower() or 'rafaqat' in name.lower() or 'masih' in name.lower():
                print(f"Salary Sheet '{sheet}' Row {r_idx+1}: '{name}' | Ref: '{row[3]}' | Pay Scale: '{row[4]}' | A/C: '{row[1]}'")
wb_sal.close()

print("\n=== Checking Attendance File ===")
wb_att = openpyxl.load_workbook('Attendanse Sheet 2026.xlsx', read_only=True)
att_names = set()
for sheet in wb_att.sheetnames:
    ws = wb_att[sheet]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if len(row) > 0 and row[0]:
            name = str(row[0]).strip()
            att_names.add(name)
            if 'rehman' in name.lower() or 'rafaqat' in name.lower() or 'masih' in name.lower():
                print(f"Attendance Sheet '{sheet}' Row {r_idx+1}: '{name}'")
wb_att.close()
