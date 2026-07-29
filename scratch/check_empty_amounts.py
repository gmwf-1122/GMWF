import openpyxl
import datetime

wb = openpyxl.load_workbook('e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx', read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n================ SHEET: {sheet_name} ================")
    
    headers = None
    first_row_idx = -1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            first_row_idx = r_idx
            break
            
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= first_row_idx:
            continue
        if not any(x is not None for x in row):
            continue
            
        receipt = row[1] if len(row) > 1 else None
        amount = row[2] if len(row) > 2 else None
        
        is_empty_or_zero = False
        if amount is None:
            is_empty_or_zero = True
        else:
            try:
                v = float(amount)
                if v == 0:
                    is_empty_or_zero = True
            except ValueError:
                # non-numeric, maybe space
                if not str(amount).strip():
                    is_empty_or_zero = True
                    
        if is_empty_or_zero:
            # Let's inspect the entire row to see if there is any other number
            other_numbers = []
            for col_idx, cell in enumerate(row):
                if col_idx in (1, 2): # Skip receipt and amount
                    continue
                if cell is None:
                    continue
                if isinstance(cell, (datetime.datetime, datetime.date)):
                    continue
                
                # Check if it contains or is a number
                cell_str = str(cell).strip()
                if not cell_str:
                    continue
                
                try:
                    num_val = float(cell_str)
                    if num_val > 0:
                        other_numbers.append((col_idx, num_val, f"col_{col_idx}: {cell}"))
                except ValueError:
                    # check if string contains something like "5K", "10000 ONLINE" etc.
                    import re
                    k_match = re.search(r'\b(\d+(?:\.\d+)?)\s*K\b', cell_str.upper())
                    if k_match:
                        other_numbers.append((col_idx, float(k_match.group(1)) * 1000, f"col_{col_idx}: {cell}"))
                    else:
                        nums = re.findall(r'\b\d+\b', cell_str)
                        for num in nums:
                            val = float(num)
                            if len(num) < 9 and val not in [2024, 2025, 2026] and val > 0:
                                other_numbers.append((col_idx, val, f"col_{col_idx}: {cell}"))
            
            if other_numbers:
                print(f"Row {r_idx+1}: Receipt={receipt}, Amount={amount}")
                print(f"  Detected other numbers: {other_numbers}")
                print(f"  Full row: {[str(x)[:20] if x is not None else None for x in row[:10]]}")

wb.close()
