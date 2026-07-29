import openpyxl
import json
from collections import defaultdict

def clean_receipt(receipt):
    if receipt is None:
        return ""
    rcpt_str = str(receipt).strip()
    if rcpt_str.endswith('.0'):
        rcpt_str = rcpt_str[:-2]
    return rcpt_str

# Load JSON records
with open('e:/GMWF/gmwf/import_donations-1.json', 'r', encoding='utf-8') as f:
    json_records = json.load(f)

json_by_exact_receipt = {r['receiptNo']: r for r in json_records}

wb = openpyxl.load_workbook('e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx', read_only=True)

# Step 1: Scan for duplicate receipts across all sheets
receipt_counts = defaultdict(int)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            break
    receipt_idx = 1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            continue
        other_fields = [row[i] for i in range(len(row)) if i != receipt_idx and row[i] is not None]
        if not other_fields:
            continue
        receipt = row[receipt_idx]
        clean_rc = clean_receipt(receipt)
        if clean_rc:
            receipt_counts[clean_rc] += 1
            
duplicate_receipts = {k for k, v in receipt_counts.items() if v > 1}
receipt_occurrences = defaultdict(int)

zeroed_rows = []
for sheet_name in ['GMWF', 'GMWF-25']:
    ws = wb[sheet_name]
    
    headers = None
    first_row_idx = -1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(x is not None for x in row):
            headers = row
            first_row_idx = r_idx
            break
            
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= first_row_idx:
            continue
        if not any(x is not None for x in row):
            continue
            
        clean_rc = clean_receipt(row[1])
        if not clean_rc:
            continue
            
        receipt_no = clean_rc
        if receipt_no in duplicate_receipts:
            receipt_occurrences[receipt_no] += 1
            occ_num = receipt_occurrences[receipt_no]
            suffix = "-a" if occ_num == 1 else "-b"
            receipt_no = f"{receipt_no}{suffix}"
            
        amount_col2 = row[2]
        
        # Check if amount_col2 is positive
        is_pos = False
        val_float = 0.0
        if amount_col2 is not None:
            try:
                val_float = float(amount_col2)
                if val_float > 0:
                    is_pos = True
            except ValueError:
                # non-numeric representation, check if whitespace
                pass
                
        if is_pos:
            if receipt_no in json_by_exact_receipt:
                json_rec = json_by_exact_receipt[receipt_no]
                json_amt = json_rec['amount']
                if json_amt != val_float:
                    zeroed_rows.append((sheet_name, r_idx+1, receipt_no, val_float, json_amt, json_rec['notes'], json_rec.get('entryType'), row))
            else:
                zeroed_rows.append((sheet_name, r_idx+1, receipt_no, val_float, None, "Not in JSON", None, row))

print(f"Total zeroed/mismatched GMWF rows: {len(zeroed_rows)}")
# Group by reason
for sheet_name, row_num, receipt_no, excel_val, json_val, notes, entry_type, row in zeroed_rows:
    print(f"[{sheet_name}] Row {row_num}: Receipt={receipt_no}, Excel={excel_val}, JSON={json_val}, EntryType={entry_type}, Notes={notes}")
    print(f"  Row details: {row[:10]}")

wb.close()
