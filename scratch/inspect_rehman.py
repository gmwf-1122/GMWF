import openpyxl

wb = openpyxl.load_workbook('GMWF  Monthly Salary.xlsx', read_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row[2] and 'rehman' in str(row[2]).lower():
            print(f"Sheet: {sheet}, Row: {r_idx+1}")
            print(f"  Data: {row[:10]}")
wb.close()
