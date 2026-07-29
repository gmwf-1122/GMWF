import openpyxl
import datetime
import re
import json
from collections import defaultdict

def clean_receipt(receipt):
    if receipt is None:
        return ""
    rcpt_str = str(receipt).strip()
    if rcpt_str.endswith('.0'):
        rcpt_str = rcpt_str[:-2]
    return rcpt_str

def parse_date_val(date_val, current_date, sheet_name):
    if date_val is None:
        return current_date
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%Y-%m-%d')
    if isinstance(date_val, datetime.date):
        return date_val.strftime('%Y-%m-%d')
    
    date_str = str(date_val).strip()
    if not date_str:
        return current_date
        
    months_map = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    
    parts = date_str.split('-')
    if len(parts) == 2:
        day_part, month_part = parts
        day_part = day_part.strip()
        month_part = month_part.strip().lower()
        
        if day_part.isdigit():
            day = int(day_part)
            month = None
            if month_part.isdigit():
                month = int(month_part)
            elif month_part[:3] in months_map:
                month = months_map[month_part[:3]]
                
            if month is not None:
                if sheet_name in ['GMWF', 'MASJID']:
                    year = 2024
                elif sheet_name == 'MASJID-25' and month == 12:
                    year = 2024
                else:
                    year = 2025
                try:
                    dt = datetime.date(year, month, day)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    pass
                    
    match_iso = re.match(r'^(\d{4})-(\d{2})-(\d{2})', date_str)
    if match_iso:
        return f"{match_iso.group(1)}-{match_iso.group(2)}-{match_iso.group(3)}"
        
    return current_date

def parse_amount_str(val, is_fallback_col=False):
    """
    Parse numeric amounts from a cell value.
    Handles: 500, 1000.0, '5K', '50K ONLINE', '10000 ONLINE', '2K ONLINE'
    Returns float or None if not parseable / looks like a year / is a date.
    
    When is_fallback_col=True (online transfer columns), we apply stricter rules:
    - Reject long integer-like values (7+ digits) that look like bank receipt numbers
    - Only accept values that clearly look like manually-typed amounts
    """
    if val is None:
        return None
    # Skip datetime objects (bank deposit dates)
    if isinstance(val, (datetime.datetime, datetime.date)):
        return None
    if isinstance(val, (int, float)):
        v = float(val)
        # Skip year-like numbers
        if is_fallback_col and 2000 <= v <= 2100:
            return None
        # In fallback columns, reject very large numbers (bank receipt numbers)
        if is_fallback_col and v >= 1_000_000:
            return None
        return v
    
    s = str(val).strip().upper()
    if not s:
        return None
    # Skip pure text like 'ONLINE', 'ONLINE TRANSFER'
    if s in ('ONLINE', 'ONLINE TRANSFER', 'ONLINE DEPOSIT'):
        return None
    
    # Try NNNk or NNN K pattern  (e.g. '5K', '50K ONLINE', '2K ONLINE')
    m = re.match(r'^([\d\.]+)\s*K\b', s)
    if m:
        return float(m.group(1)) * 1000
    
    # Try leading numeric value (e.g. '57500', '10000 ONLINE', '7000')
    m2 = re.match(r'^([\d\.]+)', s)
    if m2:
        v = float(m2.group(1))
        if is_fallback_col and 2000 <= v <= 2100:
            return None  # year-like, skip
        # In fallback columns, reject bank receipt-like numbers (7+ digits)
        if is_fallback_col and v >= 1_000_000:
            return None
        return v
    
    return None

def process_excel():
    file_path = "e:/GMWF/gmwf/GRT DAILY Donation Record.xlsx"
    wb = openpyxl.load_workbook(file_path, read_only=True)
    
    # Step 1: Scan for duplicate receipts across all sheets
    receipt_counts = defaultdict(int)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if any(x is not None for x in row):
                headers = row
                break
        
        if not headers:
            continue
            
        receipt_idx = -1
        for idx, h in enumerate(headers):
            if h and ('receipt' in str(h).lower() or 'reciept' in str(h).lower()) and 'deposit' not in str(h).lower():
                receipt_idx = idx
                break
        if receipt_idx == -1:
            receipt_idx = 1
            
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx == 0:
                continue
            other_fields = [row[i] for i in range(len(row)) if i != receipt_idx and row[i] is not None]
            if not other_fields:
                continue
                
            receipt = row[receipt_idx]
            clean_rc = clean_receipt(receipt)
            if clean_rc:
                receipt_counts[clean_rc] += 1
                
    print(f"Total distinct receipt numbers: {len(receipt_counts)}")
    duplicate_receipts = {k for k, v in receipt_counts.items() if v > 1}
    print(f"Total duplicate receipt numbers to suffix: {len(duplicate_receipts)}")
    
    # Step 2: Parse rows and construct donation records
    records = []
    receipt_occurrences = defaultdict(int)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        first_row_idx = -1
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if any(x is not None for x in row):
                headers = [str(x).strip() if x is not None else '' for x in row]
                first_row_idx = r_idx
                break
                
        if first_row_idx == -1:
            continue
            
        category_id = 'jamia' if 'masjid' in sheet_name.lower() else 'gmwf'
        
        # Hardcoded column indexes — ignore all bank deposit columns (6+)
        date_idx        = 0
        receipt_idx     = 1
        amount_idx      = 2
        # For GMWF-25 only: name, cell, receivedBy are present
        if sheet_name == 'GMWF-25':
            name_idx         = 3
            cell_idx         = 4
            received_by_idx  = 5
            # Online transfer fallback columns (data entry mistake — amount put here)
            online_fallback_cols = [6, 7]
        else:
            name_idx         = -1
            cell_idx         = -1
            received_by_idx  = 3
            online_fallback_cols = []
            
        print(f"Processing sheet: {sheet_name} (Category: {category_id})")
        
        current_date = "2024-09-29"
        sheet_count = 0
        sheet_cancelled_count = 0
        sheet_blank_skipped = 0
        sheet_online_recovered = 0
        
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx <= first_row_idx:
                continue
            
            receipt = row[receipt_idx] if receipt_idx < len(row) else None
            clean_rc = clean_receipt(receipt)
            
            # Skip completely blank rows (only check primary columns, not bank columns)
            max_col_to_check = 6 if sheet_name == 'GMWF-25' else 4
            other_fields_populated = False
            for col_i in range(min(len(row), max_col_to_check)):
                if col_i != receipt_idx and row[col_i] is not None:
                    if str(row[col_i]).strip():
                        other_fields_populated = True
                        break
            # Also check bank cols for GMWF-25 — some rows only have data there
            if not other_fields_populated and sheet_name == 'GMWF-25':
                for col_i in online_fallback_cols:
                    if col_i < len(row) and row[col_i] is not None:
                        val = row[col_i]
                        # Only count non-date values
                        if not isinstance(val, (datetime.datetime, datetime.date)):
                            if str(val).strip():
                                other_fields_populated = True
                                break
                        
            if not other_fields_populated:
                sheet_blank_skipped += 1
                continue
                
            # Date propagation
            date_val = row[date_idx] if date_idx < len(row) else None
            current_date = parse_date_val(date_val, current_date, sheet_name)
            
            # Duplicate receipt suffixing
            receipt_no = clean_rc
            if receipt_no in duplicate_receipts:
                receipt_occurrences[receipt_no] += 1
                occ_num = receipt_occurrences[receipt_no]
                suffix = "-a" if occ_num == 1 else "-b"
                receipt_no = f"{receipt_no}{suffix}"
                
            # Read donor name, cell, and received_by
            donor_name   = str(row[name_idx]).strip() if (name_idx != -1 and name_idx < len(row) and row[name_idx] is not None) else ""
            cell         = str(row[cell_idx]).strip() if (cell_idx != -1 and cell_idx < len(row) and row[cell_idx] is not None) else ""
            received_by  = str(row[received_by_idx]).strip() if (received_by_idx != -1 and received_by_idx < len(row) and row[received_by_idx] is not None) else "Excel Import"
            
            # Check if cancelled
            is_cancelled = False
            if 'cancel' in donor_name.lower():
                is_cancelled = True
                
            raw_amount = row[amount_idx] if amount_idx < len(row) else None
            if raw_amount is not None and 'cancel' in str(raw_amount).lower():
                is_cancelled = True
                
            amount_val    = 0.0
            entry_type    = 'cash'
            goods_item    = None
            payment_method = 'Cash'
            online_fallback_used = False
            
            if is_cancelled:
                amount_val = 0.0
                sheet_cancelled_count += 1
            else:
                # --- Check goods first (by donor name keywords) ---
                goods_keywords = ['pcs', 'kg', 'roti', 'quran']
                name_is_goods = any(term in donor_name.lower() for term in goods_keywords)
                
                # Try primary amount column (col2)
                amt2 = parse_amount_str(raw_amount)
                
                if amt2 is not None and amt2 > 0:
                    # Also check if the amount cell itself contains goods text
                    amt_is_goods = any(term in str(raw_amount).lower() for term in goods_keywords)
                    if name_is_goods or amt_is_goods:
                        entry_type = 'goods'
                        goods_item = donor_name if donor_name else str(raw_amount)
                        amount_val = amt2
                        payment_method = 'Goods'
                    else:
                        amount_val = amt2
                else:
                    # col2 is empty/null — check if goods by name
                    if name_is_goods:
                        entry_type = 'goods'
                        goods_item = donor_name
                        amount_val = 0.0
                        payment_method = 'Goods'
                    # Note: online transfer entries (col6 amounts) are intentionally skipped.
                    # The user's expected total only covers the Amount (col2) column.
                
                # If still 0 after all attempts → treat as cancelled
                if amount_val == 0.0 and entry_type != 'goods':
                    is_cancelled = True
                    sheet_cancelled_count += 1
                    
            # Donor anonymity
            is_anonymous = True
            donor_display_name = 'Valued Donor'
            
            if not is_cancelled and entry_type != 'goods':
                name_clean = donor_name.strip().upper()
                anonymous_names = {'UNKNOWN', 'ANONYMOUS', 'VALUED DONOR', 'ONLINE',
                                   'ONLINE TRANSFER', 'ONLINE DEPOSIT', 'SCRAP SALE', ''}
                if name_clean and name_clean not in anonymous_names:
                    is_anonymous = False
                    donor_display_name = donor_name
                
            # Build notes
            notes_parts = []
            if is_cancelled:
                notes_parts.append("CANCELLED receipt. Needs manual check.")
            elif entry_type == 'goods':
                notes_parts.append(f"In-kind donation: {goods_item}")
            if online_fallback_used:
                notes_parts.append("Online Transfer (amount from notes column)")
            if (is_cancelled or entry_type == 'goods') and donor_name:
                notes_parts.append(f"Original Text: {donor_name}")
            notes_parts.append(f"sheet: {sheet_name}.xlsx")
            notes_val = " | ".join(notes_parts)
            
            # Format phone
            phone_val = ""
            if not is_cancelled and cell:
                phone_val = re.sub(r'\D', '', cell)
                if len(phone_val) == 10 and phone_val.startswith('3'):
                    phone_val = '0' + phone_val
                    
            record = {
                'branchId':      'gujrat',
                'branchName':    'Gujrat',
                'date':          current_date,
                'amount':        amount_val,
                'receiptNo':     receipt_no,
                'donorName':     donor_display_name,
                'phone':         phone_val,
                'isAnonymous':   is_anonymous,
                'notes':         notes_val,
                'categoryId':    category_id,
                'paymentMethod': payment_method,
                'entryType':     entry_type,
                'goodsItem':     goods_item,
                'status':        'received',
                'recordedBy':    received_by if received_by else 'Excel Import'
            }
            records.append(record)
            sheet_count += 1
            
        print(f"  Finished {sheet_name}: Imported={sheet_count} "
              f"(Cancelled={sheet_cancelled_count}, OnlineRecovered={sheet_online_recovered}), "
              f"Skipped Blank={sheet_blank_skipped}")
        
    wb.close()
    
    # Summary
    gmwf_total  = sum(r['amount'] for r in records if r['categoryId'] == 'gmwf')
    jamia_total = sum(r['amount'] for r in records if r['categoryId'] == 'jamia')
    print(f"\nGMWF total:  {gmwf_total:,.0f}  (expected: 4,149,470)")
    print(f"Jamia total: {jamia_total:,.0f}")
    print(f"Grand total: {gmwf_total+jamia_total:,.0f}")
    
    output_path = "e:/GMWF/gmwf/import_donations-1.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(records, f, indent=2)
        
    print(f"\nSuccessfully generated {output_path} with {len(records)} records.")

if __name__ == "__main__":
    process_excel()
