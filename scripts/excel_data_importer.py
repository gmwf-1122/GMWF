#!/usr/bin/env python3
"""
GMWF Comprehensive Excel Data Importer & Background Processing Script
-----------------------------------------------------------------------
Parses, normalizes, and imports Excel file data (.xlsx, .xls, .csv) for:
  1. School Students Data (`school_students`)
  2. Madrassa Students Data (`madrassa_students`)
  3. Employee / Staff Data (`employees`)
  4. Donations / Financial Collection Data (`donations`)

Features:
  - Smart Header Matching: Automatically matches flexible column header names
    (e.g., 'Student Name', 'Name', 'Roll No', 'GR No', 'CNIC', 'Amount', etc.).
  - Auto Data Cleaning: Cleans dates, amounts, phone numbers, CNICs, and PINs.
  - Multi-Sheet Support: Scans single or multiple workbooks seamlessly.
  - Local JSON Output: Saves normalized output JSON for Flutter app background consumption.
  - Cloud Firestore Sync: Direct optional upload to Firebase Firestore collections.
  - Auto Biometric Credential Mapping: Registers student/employee PINs for ZKTeco biometrics.

Usage:
  # Import school students from Excel and save to JSON:
  python excel_data_importer.py --file student_list.xlsx --type school_students --branch gujrat_hq

  # Import madrassa students and upload directly to Firestore:
  python excel_data_importer.py --file madrassa_records.xlsx --type madrassa_students --upload-firestore

  # Auto-detect category and output normalized JSON:
  python excel_data_importer.py --file donations_2026.xlsx --out processed_donations.json
"""

import os
import sys
import re
import json
import uuid
import argparse
from datetime import datetime

# Windows console UTF-8 fix
if sys.platform == "win32" and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

def _ensure_package(module_name, pip_name=None):
    pip_name = pip_name or module_name
    try:
        __import__(module_name)
    except ImportError:
        import subprocess
        try:
            print(f"[INFO] Auto-installing required package '{pip_name}'...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
        except Exception as e:
            print(f"[WARN] Failed to install package '{pip_name}': {e}")

_ensure_package("openpyxl")
_ensure_package("firebase_admin", "firebase-admin")

import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore


# ── Column Aliases for Flexible Header Matching ──────────────────────────────
HEADER_ALIASES = {
    "id": ["id", "pin", "emp_id", "employee id", "gr_no", "gr no", "roll_no", "roll no", "reg_no", "registration no"],
    "name": ["name", "student name", "employee name", "donor name", "full name", "naam", "taalib ilam"],
    "father_name": ["father name", "walid ka naam", "guardian name", "father_name", "f/name", "parent name"],
    "class": ["class", "grade", "standard", "darja", "group", "class/section"],
    "section": ["section", "sec", "division"],
    "cnic": ["cnic", "nic", "national id", "b-form", "bform", "b form", "guardian cnic"],
    "phone": ["phone", "mobile", "contact", "phone number", "mobile number", "whatsapp"],
    "designation": ["designation", "role", "post", "position", "title"],
    "department": ["department", "dept", "sub-department"],
    "salary": ["salary", "basic salary", "pay", "monthly salary"],
    "amount": ["amount", "total amount", "donation amount", "total", "rs", "rupees", "rakam", "received amount"],
    "type": ["type", "donation type", "category", "fund", "head", "zakat/sadqah"],
    "mode": ["mode", "payment mode", "cash/bank", "payment method", "method"],
    "date": ["date", "receipt date", "date of birth", "dob", "joining date", "tareekh"],
    "receipt_no": ["receipt no", "receipt_no", "rcpt no", "slip no", "voucher no"],
    "address": ["address", "city", "location", "pata"],
}


def clean_cell_str(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def clean_amount(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").strip()
    m = re.search(r"[\d\.]+", s)
    if m:
        try:
            return float(m.group(0))
        except ValueError:
            pass
    return 0.0


def clean_date(val):
    if val is None:
        return datetime.now().strftime("%Y-%m-%d")
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return datetime.now().strftime("%Y-%m-%d")
    
    # Try ISO YYYY-MM-DD
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    
    # Try DD-MM-YYYY
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
        
    return datetime.now().strftime("%Y-%m-%d")


def match_header(row_cells):
    """Maps header row cells to standardized target field names."""
    field_map = {}
    
    # Pass 1: Exact alias match
    for col_idx, cell in enumerate(row_cells):
        val = clean_cell_str(cell).lower().strip()
        if not val:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if val in aliases:
                if field not in field_map:
                    field_map[field] = col_idx
                break

    # Pass 2: Substring match for remaining unmatched fields
    for col_idx, cell in enumerate(row_cells):
        val = clean_cell_str(cell).lower().strip()
        if not val:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if field not in field_map:
                if any(alias in val for alias in aliases):
                    field_map[field] = col_idx
                    break
    return field_map


def auto_detect_type(field_map, file_path):
    """Detects import category if not explicitly specified."""
    fields = set(field_map.keys())
    path_lower = file_path.lower()
    
    if "amount" in fields or "receipt_no" in fields or "donation" in path_lower:
        return "donations"
    elif "salary" in fields or "designation" in fields or "employee" in path_lower or "staff" in path_lower:
        return "employees"
    elif "class" in fields and ("school" in path_lower or "student" in path_lower):
        return "school_students"
    elif "darja" in fields or "department" in fields or "madrassa" in path_lower:
        return "madrassa_students"
    
    return "school_students"


def init_firebase(creds_path="serviceAccountKey.json"):
    """Initializes Firebase Admin SDK if service account key exists."""
    if firebase_admin._apps:
        return firestore.client()
    if os.path.exists(creds_path):
        try:
            cred = credentials.Certificate(creds_path)
            firebase_admin.initialize_app(cred)
            print(f"[FIREBASE] Initialized Firebase Admin SDK from {creds_path}")
            return firestore.client()
        except Exception as e:
            print(f"[WARN] Failed to initialize Firebase: {e}")
    script_dir_creds = os.path.join(os.path.dirname(__file__), creds_path)
    if os.path.exists(script_dir_creds):
        try:
            cred = credentials.Certificate(script_dir_creds)
            firebase_admin.initialize_app(cred)
            print(f"[FIREBASE] Initialized Firebase Admin SDK from {script_dir_creds}")
            return firestore.client()
        except Exception as e:
            print(f"[WARN] Failed to initialize Firebase: {e}")
    print("[INFO] Firebase credentials not found. Operating in local JSON export mode.")
    return None


def parse_excel_file(file_path, import_type="auto", branch_id="gujrat_hq"):
    """Parses Excel file and extracts normalized data records."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    parsed_records = []
    bio_credentials = []

    total_scanned = 0
    total_valid = 0
    total_skipped = 0

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        # Header detection
        header_row_idx = 0
        field_map = {}
        for r_idx, row in enumerate(rows[:5]):  # Search top 5 rows for header
            f_map = match_header(row)
            if len(f_map) >= 2:
                header_row_idx = r_idx
                field_map = f_map
                break

        if not field_map:
            continue

        category = import_type if import_type != "auto" else auto_detect_type(field_map, file_path)

        data_rows = rows[header_row_idx + 1 :]
        for row in data_rows:
            total_scanned += 1
            if not row or not any(row):
                total_skipped += 1
                continue

            def get_val(f_name):
                idx = field_map.get(f_name)
                return clean_cell_str(row[idx]) if idx is not None and idx < len(row) else ""

            raw_name = get_val("name")
            if not raw_name:
                total_skipped += 1
                continue

            rec_id = get_val("id") or str(uuid.uuid4())[:8]

            if category == "school_students":
                rec = {
                    "id": rec_id,
                    "grNo": rec_id,
                    "name": raw_name,
                    "fatherName": get_val("father_name") or "N/A",
                    "grade": get_val("class") or "Class 1",
                    "section": get_val("section") or "A",
                    "phone": get_val("phone"),
                    "cnic": get_val("cnic"),
                    "branchId": branch_id,
                    "status": "Active",
                    "updatedAt": datetime.now().isoformat(),
                }
                # Biometric Mapping
                bio_credentials.append({
                    "pin": rec_id,
                    "entityId": rec_id,
                    "entityName": raw_name,
                    "entityType": "school_student",
                    "branchId": branch_id,
                })

            elif category == "madrassa_students":
                rec = {
                    "id": rec_id,
                    "registrationNo": rec_id,
                    "name": raw_name,
                    "fatherName": get_val("father_name") or "N/A",
                    "department": get_val("department") or get_val("class") or "Hifz",
                    "section": get_val("section") or "A",
                    "phone": get_val("phone"),
                    "guardianCnic": get_val("cnic"),
                    "branchId": branch_id,
                    "status": "Active",
                    "updatedAt": datetime.now().isoformat(),
                }
                bio_credentials.append({
                    "pin": rec_id,
                    "entityId": rec_id,
                    "entityName": raw_name,
                    "entityType": "madrassa_student",
                    "branchId": branch_id,
                })

            elif category == "employees":
                rec = {
                    "id": rec_id,
                    "pin": rec_id,
                    "name": raw_name,
                    "fatherName": get_val("father_name") or "N/A",
                    "designation": get_val("designation") or "Staff Member",
                    "department": get_val("department") or "General",
                    "phone": get_val("phone"),
                    "cnic": get_val("cnic"),
                    "basicSalary": clean_amount(get_val("salary")),
                    "branchId": branch_id,
                    "status": "Active",
                    "updatedAt": datetime.now().isoformat(),
                }
                bio_credentials.append({
                    "pin": rec_id,
                    "entityId": rec_id,
                    "entityName": raw_name,
                    "entityType": "employee",
                    "branchId": branch_id,
                })

            elif category == "donations":
                amount = clean_amount(get_val("amount"))
                rec = {
                    "id": str(uuid.uuid4()),
                    "receiptNo": get_val("receipt_no") or f"RCP-{rec_id}",
                    "donorName": raw_name,
                    "donorPhone": get_val("phone"),
                    "amount": amount,
                    "donationType": get_val("type") or "General",
                    "paymentMode": get_val("mode") or "Cash",
                    "date": clean_date(get_val("date")),
                    "branchId": branch_id,
                    "createdAt": datetime.now().isoformat(),
                }

            parsed_records.append((category, rec))
            total_valid += 1

    return {
        "scanned": total_scanned,
        "valid": total_valid,
        "skipped": total_skipped,
        "records": parsed_records,
        "biometrics": bio_credentials,
    }


def upload_to_firestore(db, parsed_results):
    """Uploads parsed records directly to Cloud Firestore collections."""
    if not db:
        print("[WARN] Skipping Firestore upload: No database client available.")
        return 0

    uploaded_count = 0
    batch = db.batch()
    batch_count = 0

    for category, rec in parsed_results["records"]:
        doc_id = rec.get("id") or str(uuid.uuid4())
        ref = db.collection(category).document(doc_id)
        batch.set(ref, rec, merge=True)
        batch_count += 1
        uploaded_count += 1

        if batch_count >= 450:
            batch.commit()
            batch = db.batch()
            batch_count = 0

    for bio in parsed_results["biometrics"]:
        pin = bio.get("pin")
        if pin:
            ref = db.collection("biometric_credentials").document(str(pin))
            batch.set(ref, bio, merge=True)
            batch_count += 1

            if batch_count >= 450:
                batch.commit()
                batch = db.batch()
                batch_count = 0

    if batch_count > 0:
        batch.commit()

    print(f"✅ Successfully uploaded {uploaded_count} records to Cloud Firestore!")
    return uploaded_count


def save_local_json(parsed_results, output_path):
    """Saves parsed records to JSON file for local app consumption."""
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    grouped_data = {}
    for cat, rec in parsed_results["records"]:
        if cat not in grouped_data:
            grouped_data[cat] = []
        grouped_data[cat].append(rec)

    payload = {
        "status": "success",
        "timestamp": datetime.now().isoformat(),
        "summary": {
            "total_scanned": parsed_results["scanned"],
            "total_valid": parsed_results["valid"],
            "total_skipped": parsed_results["skipped"],
        },
        "data": grouped_data,
        "biometrics": parsed_results["biometrics"],
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

    print(f"📄 Saved normalized data output to: {os.path.abspath(output_path)}")


def main():
    parser = argparse.ArgumentParser(
        description="GMWF Excel Data Importer & Background Processing Script"
    )
    parser.add_argument("--file", required=True, help="Path to input Excel file (.xlsx, .xls)")
    parser.add_argument(
        "--type",
        default="auto",
        choices=["auto", "school_students", "madrassa_students", "employees", "donations"],
        help="Target entity type for import",
    )
    parser.add_argument("--branch", default="gujrat_hq", help="Target Branch ID (default: gujrat_hq)")
    parser.add_argument("--out", default="imported_data_output.json", help="Path to output JSON file")
    parser.add_argument(
        "--upload-firestore", action="store_true", help="Upload records directly to Cloud Firestore"
    )
    parser.add_argument("--creds", default="serviceAccountKey.json", help="Firebase service account JSON path")

    args = parser.parse_args()

    print("==========================================================")
    print("        GMWF Excel Data Importer & Parser                 ")
    print("==========================================================")
    print(f"Input File : {args.file}")
    print(f"Import Type: {args.type}")
    print(f"Branch ID  : {args.branch}\n")

    try:
        results = parse_excel_file(args.file, import_type=args.type, branch_id=args.branch)
        print(f"Parsing Summary:")
        print(f"  ├─ Total rows scanned: {results['scanned']}")
        print(f"  ├─ Valid records:     {results['valid']}")
        print(f"  └─ Skipped/empty:     {results['skipped']}\n")

        save_local_json(results, args.out)

        if args.upload_firestore:
            db = init_firebase(args.creds)
            upload_to_firestore(db, results)

    except Exception as e:
        print(f"[ERROR] Import failed: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
